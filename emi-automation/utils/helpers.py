"""EMI formula, UI parsing, and test data loading."""

from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any, TypedDict

from openpyxl import load_workbook

TESTDATA_DIR = Path(__file__).resolve().parent.parent / "testdata"


class EmiTotals(TypedDict):
    emi: int
    total_payment: int
    total_interest: int


def parse_currency(value: str) -> int:
    """Convert '₹44,986' or '50,00,000' to integer rupees."""
    cleaned = re.sub(r"[^\d.]", "", value.strip())
    if not cleaned:
        return 0
    return int(float(cleaned))


def _exact_emi(principal: float, annual_rate_pct: float, tenure_months: int) -> float:
    if tenure_months <= 0 or principal <= 0:
        return 0.0
    if annual_rate_pct == 0:
        return principal / tenure_months

    monthly_rate = annual_rate_pct / 12 / 100
    factor = math.pow(1 + monthly_rate, tenure_months)
    return principal * monthly_rate * factor / (factor - 1)


def calculate_totals(
    principal: float, annual_rate_pct: float, tenure_months: int
) -> EmiTotals:
    """Return EMI, total payment, and total interest (all rounded to rupees)."""
    exact_emi = _exact_emi(principal, annual_rate_pct, tenure_months)
    emi = round(exact_emi)
    total_payment = round(exact_emi * tenure_months)
    total_interest = total_payment - round(principal)
    return {
        "emi": emi,
        "total_payment": total_payment,
        "total_interest": total_interest,
    }


def get_test_data(key: str, file_name: str = "test_cases.json") -> Any:
    """Return a top-level key from test data JSON."""
    file_path = TESTDATA_DIR / file_name
    if not file_path.exists():
        raise FileNotFoundError(f"Test data file not found: {file_path}")

    with file_path.open(encoding="utf-8") as handle:
        data = json.load(handle)

    if key not in data:
        raise KeyError(f"Key '{key}' not found in {file_name}. Available: {list(data)}")
    return data[key]


def parse_excel_export(file_path: Path) -> dict[str, int | float | list[dict]]:
    """Read loan summary and amortization rows from a downloaded EMI Excel file."""
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    labels: dict[str, int | float] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2, values_only=True):
        label, value = row[0], row[1] if len(row) > 1 else None
        if label and value is not None:
            labels[str(label).strip()] = value

    amount_label = next(key for key in labels if key.endswith("Loan Amount"))
    summary = {
        "principal": int(round(float(labels[amount_label]))),
        "rate": float(labels["Interest Rate (%)"]),
        "tenure_months": int(round(float(labels["Loan Tenure (months)"]))),
        "emi": int(round(float(labels["Loan EMI"]))),
        "total_interest": int(round(float(labels["Total Interest Payable"]))),
        "total_payment": int(round(float(labels["Total Payment (Principal + Interest)"]))),
    }

    schedule_rows: list[dict] = []
    header_row = None
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True), start=1
    ):
        if row[0] == "Month #":
            header_row = row_index
            break

    if header_row:
        for row in worksheet.iter_rows(
            min_row=header_row + 1, max_row=worksheet.max_row, values_only=True
        ):
            if row[0] is None:
                continue
            schedule_rows.append(
                {
                    "month": int(row[0]),
                    "principal": int(round(float(row[2]))),
                    "interest": int(round(float(row[3]))),
                    "total_payment": int(round(float(row[4]))),
                    "balance": int(round(float(row[5]))),
                }
            )

    return {**summary, "schedule_rows": schedule_rows}
